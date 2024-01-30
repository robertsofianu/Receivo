import json
import os
import sqlite3
import shutil
import openpyxl
import requests
import zipfile

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from twilio.rest import Client

# DataBase Path
DB_PATH = r'db\produse.db'

# Initialize path variables
JSON_PATH_NIR_NUMBER = r'JsonFiles\NirNumber.json'
JSON_PATH_EXCEL_NUMBER = r'JsonFiles\ExcelNumber.json'
JSON_PATH_ROW_NUMBER = r'JsonFiles\NirNumber.json'
JSON_PATH_FURIZORI = r'JsonFiles\Furnizori.json'
JSON_PATH_PRODUSE = r'JsonFiles\AllProducts.json'
JSON_PATH_CUSTOMERS = r'JsonFiles\customers.json'
JSON_PATH_DICT_ALL_DATA = r'JsonFiles\dict_all_product.json'
PATH_TO_USERS = "your\\path\\to\\users"

# Initialize bool variables
bool_data_complete = True
bool_is_in_db = False
is_complete = True
invalid_sign_up = False
invaid_login = False
emailwassent = False
passwords_match = True
bool_cant = False
modify_product = ''
succes_mod = False
all_products = []
bool_invalid_desk_data = False




def fct_path(user, nr, furnizor):
    data_separated = nr.split('-')
    months = [  
        '1_January',
        '2_February',
        '3_March',
        '4_April',
        '5_May',
        '6_June',
        '7_July',
        '8_August',
        '9_September',
        '10_October',
        '11_November',
        '12_December'
        ]
    file1 = f'{PATH_TO_USERS}\\{user}\\{data_separated[0]}\\{months[int(data_separated[1]) - 1]}\\{nr}[{furnizor}].xlsx'
    return file1

def fct_path_list_product(user):
    file =  f'{PATH_TO_USERS}\\{user}\\list_nir_products.json'
    return file


def fct_path_table(user):
    file =  f'{PATH_TO_USERS}\\{user}\\table_final_nr.json'
    return file

def fct_path_dict_all_product(user):
    file =  f'{PATH_TO_USERS}\\{user}\\dict_all_product.json'
    return file

def fct_path_ExcelNumber(user):
    file =  f'{PATH_TO_USERS}\\{user}\\ExcelNumber.json'
    return file

def fct_path_NumberRowExcel(user):
    file =  f'{PATH_TO_USERS}\\{user}\\NumberRowExcel.json'
    return file

def fct_path_product_modify_name(user):
    file = f'{PATH_TO_USERS}\\{user}\\product_modify_name.json'
    return file
    
def fct_path_bool_is_home(user):
    file = f'{PATH_TO_USERS}\\{user}\\json_is_hame_bool.json'
    return file


def fct_get_fine_data(l: list):
    for tuple in l:
        for result in tuple:
            return result

 
def fct_read_json(jsonPath):  
    with open(jsonPath, 'r') as file:
        json_content = file.read()
    data = json.loads(json_content)
    return data[-1]

def fct_read_json_all_d(jsonPath):  
    with open(jsonPath, 'r') as file:
        json_content = file.read()
    data = json.loads(json_content)
    return data


def fct_increment_nir_nr(jsonPath):
    last_nir = int(fct_read_json(jsonPath))
    new_nir = [last_nir + 1]
    with open(jsonPath, 'w') as jf:
        json.dump(new_nir, jf, indent=4)


def fct_write_JSON(jsonPath, file1):
    file = [
        file1
    ]
    with open(jsonPath, 'w') as jf:
        json.dump(file, jf, indent=4)




def fct_insert_db(produs, tva, u_m, pf, pv):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
    INSERT INTO preturi (produs, TVA, u_m, pret_furnizor, pret_final) VALUES (?, ?, ?, ?, ?)
    """, (produs, tva, u_m, pf, pv,))
    con.commit()


def fct_create_excel_file(data, nrFac, furnizor, delegat, user):
    JSON_PATH_EXCEL_NUMBER = fct_path_ExcelNumber(user)

    nrNir = 2222
    fct_append_json(data, user)
    nr_excel = fct_return_data_name(user)
    
    file1 = fct_path(user, nr_excel, furnizor)

    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.row_dimensions[8].height = 75
    worksheet.row_dimensions[1].height = 27

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 5
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['I'].width = 10

    cell_d1 = worksheet['D1']
    cell_d1.value = "NOTA DE RECEPTIE"
    font_style = Font(size=18, bold=True)
    cell_d1.font = font_style

    worksheet['A1'] = 'S.C.'
    worksheet['B1'] = 'PROTOSOF S.R.L.'
    worksheet['H1'] = 'Nr.'
    worksheet['J1'] = 'Din'
    worksheet['K1'] = data

    worksheet['B2'] = 'Protopopesti'
    worksheet['A2'] = 'Localitatea'

    worksheet['D3'] = 'Furnzor'
    worksheet['D4'] = 'Nr. si data facturii sau avizului de expediere'
    worksheet['D5'] = 'Numele si prenumele delegatului'
    worksheet['D6'] = 'B.l. Seria …...... nr. ….......... eliberat de Politia ...................................'

    worksheet['I1'] = nrNir
    worksheet['I3'] = furnizor
    worksheet['I4'] = nrFac
    worksheet['I5'] = delegat

    cell = worksheet['I1']
    cell.value = nrNir
    cell.alignment = Alignment(
        horizontal='left', vertical='bottom')

    cell = worksheet['A8']
    cell.value = "Denumire Produs"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['B8']
    cell.value = "U/M"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['C8']
    cell.value = "Dupa factura sau aviz"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['D8']
    cell.value = "La receptie"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['E8']
    cell.value = "Pret Furnizor"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['F8']
    cell.value = "Valoare furnizor col. 3 x 4"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['G8']
    cell.value = "T.V.A. Incasat de furnizor"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['H8']
    cell.value = "Total valoare factura col. 5 + 6"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['I8']
    cell.value = "Pret cu adasul practicat %"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['J8']
    cell.value = "Valoare la pret cu adaos col. 3 x 8"
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['K8']
    cell.value = "Pret cu adaos si T.V.A."
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = worksheet['L8']
    cell.value = "Valoare la pret cu adaos si T.V.A."
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')


    workbook.save(file1)
    workbook.close()


def fct_get_id_bool_elem_from_db(product):
    response = ''
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    cur.execute("SELECT id FROM preturi WHERE produs = ?", (product,))

    ids = cur.fetchall()
    if ids:
        response = 1
    else:
        response = -1
    
    return response

def fct_get_id_elem_from_db(product):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    cur.execute("SELECT id FROM preturi WHERE produs = ?", (product,))

    ids = cur.fetchall()
    id = fct_get_fine_data(ids)
    
    return id


def fct_get_data_db(prod):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    cur.execute('SELECT TVA FROM preturi WHERE produs = ?', (prod,))
    tva = cur.fetchall()
    fine_tva = fct_get_fine_data(tva)

    cur.execute('SELECT u_m FROM preturi WHERE produs = ?', (prod,))
    u_m = cur.fetchall()
    fine_um = fct_get_fine_data(u_m)

    cur.execute('SELECT pret_furnizor FROM preturi WHERE produs = ?', (prod,))
    pret_furnizor = cur.fetchall()
    fine_pret_furnizor = fct_get_fine_data(pret_furnizor)

    cur.execute('SELECT pret_final FROM preturi WHERE produs = ?', (prod,))
    pret_final = cur.fetchall()
    fine_pret_final = fct_get_fine_data(pret_final)

    return prod, fine_tva, fine_um, fine_pret_furnizor, fine_pret_final


def fct_create_excel_tables(t: list, cantitate, user, furnizor):
    JSON_PATH_ROW_NUMBER = fct_path_NumberRowExcel(user)
    att = fct_read_json(JSON_PATH_ROW_NUMBER)
    a = fct_read_json
    nr_excel = fct_return_data_name(user)
    excel_path = fct_path(user, nr_excel, furnizor=furnizor)

    wb = load_workbook(excel_path)
    ws = wb.active

    if t[1] == '9':
        tva_f = float('0.09')
    else:
        tva_f = float('0.19')

    cell = ws[f'A{att}']
    cell.value = t[0]
    cell.alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')

    cell = ws[f'B{att}']
    cell.value = t[2]
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'C{att}']
    cell.value = float(cantitate)
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'D{att}']
    cell.value = float(cantitate)
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'E{att}']
    cell.value = float(t[3])
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'F{att}']
    cell.value = f'=PRODUCT(D{att},E{att})'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'G{att}']
    cell.value = f'=F{att}*{tva_f}'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'I{att}']
    cell.value = float(t[4])
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'L{att}']
    cell.value = f'=PRODUCT(D{att},I{att})'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')
    fct_increment_nir_nr(JSON_PATH_ROW_NUMBER)

    wb.save(excel_path)


def fct_reset_excel_row(user):
    JSON_PATH_ROW_NUMBER = fct_path_NumberRowExcel(user)
    with open(JSON_PATH_ROW_NUMBER, 'w') as jf:
        json.dump([9], jf, indent=4)


def fct_terminare(user, furnizor):
    nr_excel = fct_return_data_name(user)
    excel_path = fct_path(user, nr_excel, furnizor=furnizor)
    wb = load_workbook(excel_path)
    ws = wb.active
    JSON_PATH_ROW_NUMBER = fct_path_NumberRowExcel(user)

    att = fct_read_json(JSON_PATH_ROW_NUMBER)

    table_range = f"A8:L{att + 1}"
    table = Table(displayName="MyTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleLight15", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    cell = ws[f'A{att + 1}']
    cell.value = 'Total'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'F{att + 1}']
    cell.value = f'=SUM(F9:F{att})'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'G{att + 1}']
    cell.value = f'=SUM(G9:G{att})'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    cell = ws[f'L{att + 1}']
    cell.value = f'=SUM(L9:L{att})'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    ws.add_table(table)

    row_number = att + 2

    colums = 'ABCDEFGHIJKL'

    for col in colums:
        if col == 'A' or col == 'H':
            cell = ws[f'{col}{row_number}']
            border = Border(left=Side(border_style="thin"),
                            top=Side(border_style="thin"))

        elif col == 'E' or col == 'L':
            cell = ws[f'{col}{row_number}']
            border = Border(right=Side(border_style="thin"),
                            top=Side(border_style="thin"))
        else:
            cell = ws[f'{col}{row_number}']
            border = Border(top=Side(border_style="thin"))
        cell.border = border

    ws[f'A{row_number}'] = """
    Numele si prenumele receptionerilor: 
    """

    ws[f'F{row_number}'] = """
    Semnatura
    """

    ws[f'H{row_number}'] = """
    Val. Marfa furnizor (col. 5)
    """

    cell = ws[f'L{row_number}']
    cell.value = f'=SUM(F{att+1}:G{att+1})'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    row_number += 1
    for col in colums:
        if col == 'A' or col == 'H':
            cell = ws[f'{col}{row_number}']
            border = Border(left=Side(border_style="thin"))

        elif col == 'E' or col == 'L':
            cell = ws[f'{col}{row_number}']
            border = Border(right=Side(border_style="thin"))
        else:
            continue
        cell.border = border
    ws[f'H{row_number}'] = """
    Val. Ambalaj furnizor
    """

    row_number += 1
    for col in colums:
        if col == 'A' or col == 'H':
            cell = ws[f'{col}{row_number}']
            border = Border(left=Side(border_style="thin"))

        elif col == 'E' or col == 'L':
            cell = ws[f'{col}{row_number}']
            border = Border(right=Side(border_style="thin"))
        else:
            continue
        cell.border = border
    ws[f'H{row_number}'] = """
    Val. Adaos comercial (col. 9-5)
    """

    sign_mama = r'singnitures\semnatura mama.png'
    sign_tata = r'singnitures\semnatura tata.png'

    img1 = Image(sign_mama)
    img2 = Image(sign_tata)

    img1.width = 80  
    img1.height = 50

    img2.width = 80  
    img2.height = 50

    ws.add_image(img1, f"F{row_number}")
    ws.add_image(img2, f"C{row_number}")



    row_number += 1
    for col in colums:
        if col == 'A' or col == 'H':
            cell = ws[f'{col}{row_number}']
            border = Border(left=Side(border_style="thin"))

        elif col == 'E' or col == 'L':
            cell = ws[f'{col}{row_number}']
            border = Border(right=Side(border_style="thin"))
        else:
            continue
        cell.border = border
    ws[f'H{row_number}'] = """
    Val. T.V.A. Neexigibila (col. 11-9)
    """

    row_number += 1
    for col in colums:
        if col == 'A' or col == 'H':
            cell = ws[f'{col}{row_number}']
            border = Border(left=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))

        elif col == 'E' or col == 'L':
            cell = ws[f'{col}{row_number}']
            border = Border(right=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))
        else:
            cell = ws[f'{col}{row_number}']
            border = Border(bottom=Side(border_style="thin"))
        cell.border = border
    ws[f'H{row_number}'] = """
    TOTAL VALOARE AMANUNT
    """

    cell = ws[f'L{row_number}']
    cell.value = f'=SUM(L9:L{row_number-6})'
    cell.number_format = '0.00'
    cell.alignment = Alignment(
        horizontal='center', vertical='center')

    wb.save(excel_path)


    JSON_PATH_EXCEL_NUMBER = fct_path_ExcelNumber(user)


    fct_reset_excel_row(user)
    fct_increment_nir_nr(JSON_PATH_NIR_NUMBER)
    nr_excel = fct_read_json(JSON_PATH_EXCEL_NUMBER)
    excel_path = f'path\\to\\root\\folder\\{user}\\Book{nr_excel}.xlsx'



def fct_append_list(nr: int, item: str):
    if nr == 1: #produse
        l = fct_read_json(JSON_PATH_PRODUSE)
        jsonpath = JSON_PATH_PRODUSE
    elif nr == 2: #furnizori
        l = fct_read_json(JSON_PATH_FURIZORI)
        jsonpath = JSON_PATH_FURIZORI

    if item not in l:
        item = item.upper()
        l.append(item)
        fct_write_JSON(file1=l, jsonPath=jsonpath)


def fct_update_product_info(product: str, tva: str, pp:float, sp: float):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute('UPDATE preturi SET TVA = ? WHERE produs = ? ', (tva, product))
    con.commit()

    cur.execute('UPDATE preturi SET pret_furnizor = ? WHERE produs = ? ', (pp,product,))
    con.commit()

    cur.execute('UPDATE preturi SET pret_final = ? WHERE produs = ? ', (sp, product,))
    con.commit()

def fct_update_product_info_id(product: str, tva: str, pp:float, sp: float, id: int):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute('UPDATE preturi SET produs = ? WHERE id = ? ', (product, id,))
    con.commit()

    cur.execute('UPDATE preturi SET TVA = ? WHERE id = ? ', (tva, id,))
    con.commit()

    cur.execute('UPDATE preturi SET pret_furnizor = ? WHERE id = ? ', (pp, id,))
    con.commit()

    cur.execute('UPDATE preturi SET pret_final = ? WHERE id = ? ', (sp, id,))
    con.commit()

def is_integer(s: str):
    try:
        int(s)
        return True
    except ValueError:
        return False

def is_float(s: str):
    try:
        if ',' in s:
            s = list(s)
            ind = s.index(',')
            s[ind] = '.'
            s = ''.join(s)
            
        float(s)
        return True, s
    except ValueError:
        return False

def fct_write_formated_price(p):
    if is_integer(p) == True:
        r = f'{p}.00'

    elif is_float(p)[0] == True:
        
        p = is_float(p)[1].split('.')

        if len(p[1]) == 1:
            r =  f'{p[0]}.{p[1]}0'

        elif len(p[1]) > 2:
            r = f'{p[0]}.{p[1][:2]}'
        
        else:
            r = f'{p[0]}.{p[1]}'

    return r



def fct_append_json(a, user):
    json_path = fct_path_ExcelNumber(user)
    with open(json_path, 'r') as file:
        content = file.read()
    data = json.loads(content)
    data.append(a)
    with open(json_path, 'w') as file:
        json.dump(data, file, indent=4)


def fct_return_data_name( user):
    json_path = fct_path_ExcelNumber(user)

    with open(json_path, 'r') as file:
        content = file.read()
    data = json.loads(content)
    c_data = data.count(data[-1])
    return f"{data[-1]}({c_data})"

def fct_validate_new_users(user, email):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    cur.execute('SELECT id FROM temporary_users WHERE username = ?', (user, ))
    rasp_user = cur.fetchall()

    cur.execute('SELECT id FROM temporary_users WHERE email = ?', (email, ))
    rasp_email = cur.fetchall()

    if not rasp_user and not rasp_email:
        cur.execute('INSERT INTO temporary_users (username, email) VALUES (?, ?)', (user, email,))
        con.commit()
        return 1 # available user
    else:
        return -1 # user or email unvailable
    

def modify_product_LIST(item, item2, user):
    j_P = JSON_PATH_PRODUSE
    list_all_products = fct_read_json(j_P)
    if item in list_all_products:
        index_item = list_all_products.index(item)
        list_all_products[index_item] = item2
        fct_write_JSON(j_P ,list_all_products)

        json_mod = fct_path_product_modify_name(user)
        data = fct_read_json_all_d(json_mod)
        data[1] = item2
        with open(json_mod, 'w') as file:
            json.dump(data, file, indent = 4)


    return 0



def fct_customers_list():
    data = fct_read_json_all_d(JSON_PATH_CUSTOMERS)
    list_all_complete_names = []

    for element in data:
        list_all_complete_names.append(element[0])
        
    list_all_complete_names = set(list_all_complete_names)
    return list_all_complete_names



def fct_sent_sms(date, user, amount, info_text, customer):
    account_sid = os.getenv('TWILIO_SMS_PROVIDER_SID')
    auth_token = os.getenv('TWILIO_SMS_PROVIDER_TOKEN')
    client = Client(account_sid, auth_token)

    message = client.messages.create(
    from_='+16179347243',
    body=f"""
NEW TRANSACTION
Customer: {customer}
User: {user}
Date: {date}
Amount: {amount} RON
Other Info: {info_text}
    """,
    to='+40785722833'
    )



def fct_trasfer_files(user, data):
    if data == None:
        return 0 # Do not execute
    else:
        months = [  
            '1_January',
            '2_February',
            '3_March',
            '4_April',
            '5_May',
            '6_June',
            '7_July',
            '8_August',
            '9_September',
            '10_October',
            '11_November',
            '12_December'
            ]
        
        data = data.split("-")

        source_folder = f"{PATH_TO_USERS}\\{user}\\{data[0]}\\{months[int(data[1]) - 1]}"
        destination_folder = f"path\\to\\your\\destination\\folder\\{data[0]}\\{months[int(data[1]) - 1]}"

        
        try:
            
            os.makedirs(destination_folder, exist_ok=True)

            # Iterate through files in the source folder
            for filename in os.listdir(source_folder):
                source_file_path = os.path.join(source_folder, filename)
                destination_file_path = os.path.join(destination_folder, filename)

                shutil.copy2(source_file_path, destination_file_path)
                return 1 # Success
        except Exception as e:
            print(f"An error occurred: {e}")
            return 0
        

def fct_download_to_local(date):
    date_split = date.split('-')
    months = [  
        '1_January',
        '2_February',
        '3_March',
        '4_April',
        '5_May',
        '6_June',
        '7_July',
        '8_August',
        '9_September',
        '10_October',
        '11_November',
        '12_December'
    ]
    year = date_split[0]
    month = months[int(date_split[1]) - 1]
    username = os.getenv('POYTHON_ANYWHERE_USERNAME')
    token = os.getenv('POYTHON_ANYWHERE_TOKEN')
    path = f'path/to/python_anywher/folder'

    response = requests.get(
        f'https://www.pythonanywhere.com/api/v0/user/{username}/files/path{path}'.format(
            username=username
        ),
        headers={'Authorization': 'Token {token}'.format(token=token)}
    )
    download_links = []

    if response.status_code == 200:

        data = response.content
        data_dict = json.loads(data.decode('utf-8'))
        print(data_dict)
        for key, value in data_dict.items():
            download_links.append(value["url"])
        

    for down_link in download_links:
        down_link_split = down_link.split('/')
        year = down_link_split[-3]
        month = down_link_split[-2]
        fileName = down_link_split[-1]
        

        destination = f"path\\to\\your\\destination\\folder\\{year}\\{month}\\{fileName}"
        try:
            os.mkdir(f"path\\to\\your\\destination\\folder\\{year}")
            os.mkdir(f"path\\to\\your\\destination\\folder\\{year}\\{month}")
        except FileExistsError:
            print("Error")
        try:
            os.mkdir(f"path\\to\\your\\destination\\folder\\{year}\\{month}")
        except FileExistsError:
            print("Error")


        token = os.getenv('POYTHON_ANYWHERE_TOKEN')
        response = requests.get(
            down_link,
            headers={'Authorization': 'Token {token}'.format(token=token)}
        )
        with open(destination, 'wb') as file:
            file.write(response.content)
        

if __name__ == '__main__':
    print()
